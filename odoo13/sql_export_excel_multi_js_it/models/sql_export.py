from odoo import api, exceptions, fields, models, _
from io import BytesIO
import base64
import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except:
    install('openpyxl==3.0.5')

class SqlExport(models.Model):
    _inherit = 'sql.export'
    parent_id = fields.Many2one('sql.export', 'SQL Principal')
    additional_pages = fields.One2many('sql.export', 'parent_id', string="Paginas Adicionales")

    def excel_get_data_from_query(self, variable_dict):
        #raise ValueError(str(variable_dict))
        self.ensure_one()
        res = self._execute_sql_request(
            params=variable_dict, mode='fetchall', header=self.header)
        #raise ValueError(res)
        # Case we insert data in an existing excel file.
        pag_add = []
        if self.attachment_id:
            datas = self.attachment_id.datas
            infile = BytesIO()
            infile.write(base64.b64decode(datas))
            infile.seek(0)
            wb = openpyxl.load_workbook(filename=infile)
            sheets = wb.worksheets
            try:
                ws = sheets[self.sheet_position - 1]
            except IndexError:
                raise exceptions.ValidationError(
                    _("The Excel Template file contains less than %s sheets "
                      "Please, adjust the Sheet Position parameter."))
            row_position = self.row_position or 1
            col_position = self.col_position or 1
        # Case of excel file creation
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.name
            if self.additional_pages:
                for p in self.additional_pages:
                    pag_add.append(p)

        row_position = 1
        col_position = 1
        for index, row in enumerate(res, row_position):
            for col, val in enumerate(row, col_position):
                ws.cell(row=index, column=col).value = val

        #crear las demas paginas si tie ne
        if pag_add:
            for p in pag_add:
                self2 = self.env['sql.export'].search([('id','=',p.id)])
                #raise ValueError(self2)
                resx = self2._execute_sql_request(
                    params=variable_dict, mode='fetchall', header=self.header)
                wsx = wb.create_sheet(str(p.name))
                #raise ValueError(resx)
                for index, row in enumerate(resx, row_position):
                    for col, val in enumerate(row, col_position):
                        try:
                            wsx.cell(row=index, column=col).value = val
                        except:
                            a = 1

        output = BytesIO()
        wb.save(output)
        output.getvalue()
        output_datas = base64.b64encode(output.getvalue())
        output.close()
        return output_datas