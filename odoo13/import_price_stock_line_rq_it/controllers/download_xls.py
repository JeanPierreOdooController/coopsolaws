# -*- coding: utf-8 -*-

from odoo import http
from odoo.http import request
from odoo.addons.web.controllers.main import content_disposition
import base64
import openpyxl  # Agrega esta importación
from openpyxl import load_workbook
from io import BytesIO

class Download_xls(http.Controller):

	@http.route('/web/binary/download_template_update_price_stock_picking_id/<int:picking_id>', type='http', auth="public")
	def download_template_move_line(self, picking_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_price_stock_moves.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Actualización de Transferencia.xlsx'
		filecontent = base64.b64decode(filecontent)

		workbook = openpyxl.load_workbook(BytesIO(filecontent))

		sheet = workbook['IMPORTADOR']
		picking = request.env['stock.picking'].sudo().search([('id', '=', picking_id)])
		row = 2
		for i in picking.move_ids_without_package:
			for j in i.move_line_ids:
				sheet[f'A{row}'] = j.id
				sheet[f'B{row}'] = j.product_id.default_code
				sheet[f'C{row}'] = j.product_id.name
				# sheet[f'D{row}'] = picking.move_line_ids_without_package.filtered(lambda l: l.product_id.id == i.product_id.id).lot_id.name
				sheet[f'D{row}'] = j.lot_id.name
				sheet[f'E{row}'] = j.qty_done
				sheet[f'F{row}'] = i.price_unit_it
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])


	@http.route('/web/binary/download_template_update_price_stock_moves/<stock_move_ids>', type='http', auth="public")
	def download_template_move_line(self, stock_move_ids, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_price_stock_moves_price.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Actualización de Movimientos.xlsx'
		filecontent = base64.b64decode(filecontent)

		workbook = openpyxl.load_workbook(BytesIO(filecontent))
		stock_moves = [int(id.strip()) for id in stock_move_ids[1:-1].split(',') if id.strip().isdigit()]
		sheet = workbook['IMPORTADOR']
		moves = request.env['stock.move'].sudo().search([('id', 'in', stock_moves)])
		row = 2		
		for j in moves:
				sheet[f'A{row}'] = j.id
				sheet[f'B{row}'] = j.price_unit_it
				sheet[f'C{row}'] = "[%s] %s"%(j.product_id.default_code,j.product_id.name)				
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])