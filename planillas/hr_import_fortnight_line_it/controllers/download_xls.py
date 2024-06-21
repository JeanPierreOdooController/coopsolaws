# -*- coding: utf-8 -*-

from odoo import http
from odoo.http import request
from odoo.addons.web.controllers.main import content_disposition
import base64
import openpyxl  # Agrega esta importación
from openpyxl import load_workbook
from io import BytesIO

class Download_xls(http.Controller):

	@http.route('/web/binary/download_template_update_income_fortnight_lines/<int:fortnight_id>', type='http', auth="public")
	def download_template_update_income_fortnight_lines(self, fortnight_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_fortnight_lines.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Importador Ingreso de Lineas Quincenales.xlsx'
		filecontent = base64.b64decode(filecontent)
		workbook = openpyxl.load_workbook(BytesIO(filecontent))
		sheet = workbook['IMPORTADOR']
		fortnight = request.env['hr.quincenales'].sudo().search([('id', '=', fortnight_id)])
		row = 2
		for i in fortnight.quincenales_lines:
			if i.quincenales_ingresos_lines:
				for income in i.quincenales_ingresos_lines:		
					sheet[f'A{row}'] = i.codigo_trabajador
					sheet[f'B{row}'] = income.id
					sheet[f'C{row}'] = i.employee_id.name
					sheet[f'D{row}'] = income.concepto_id.code
					sheet[f'E{row}'] = income.monto
					sheet[f'F{row}'] = 'Ingreso'
					row += 1
			else:
				sheet[f'A{row}'] = i.codigo_trabajador
				sheet[f'C{row}'] = i.employee_id.name
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])

	@http.route('/web/binary/download_template_update_discounts_fortnight_lines/<int:fortnight_id>', type='http', auth="public")
	def download_template_update_discounts_fortnight_lines(self, fortnight_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_fortnight_lines.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Importador Descuento de Lineas Quincenales.xlsx'
		filecontent = base64.b64decode(filecontent)
		workbook = openpyxl.load_workbook(BytesIO(filecontent))
		sheet = workbook['IMPORTADOR']
		fortnight = request.env['hr.quincenales'].sudo().search([('id', '=', fortnight_id)])
		row = 2
		for i in fortnight.quincenales_lines:
			if i.quincenales_descuentos_lines:
				for discounts in i.quincenales_descuentos_lines:		
					sheet[f'A{row}'] = i.codigo_trabajador
					sheet[f'B{row}'] = discounts.id
					sheet[f'C{row}'] = i.employee_id.name
					sheet[f'D{row}'] = discounts.concepto_id.code
					sheet[f'E{row}'] = discounts.monto
					sheet[f'F{row}'] = 'Descuento'
					row += 1
			else:
				sheet[f'A{row}'] = i.codigo_trabajador
				sheet[f'C{row}'] = i.employee_id.name
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])
	
	@http.route('/web/binary/download_template_out_of_fortnight_update_fortnight_lines/<int:fortnight_id>', type='http', auth="public")
	def download_template_out_of_fortnight_update_fortnight_lines(self, fortnight_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_fortnight_lines.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Importador Fuera de Quincena de Lineas Quincenales.xlsx'
		filecontent = base64.b64decode(filecontent)
		workbook = openpyxl.load_workbook(BytesIO(filecontent))
		sheet = workbook['IMPORTADOR']
		fortnight = request.env['hr.quincenales'].sudo().search([('id', '=', fortnight_id)])
		row = 2
		for i in fortnight.quincenales_lines:
			if i.quincenales_conceptos_lines:
				for out_of_fortnight in i.quincenales_conceptos_lines:		
					sheet[f'A{row}'] = i.codigo_trabajador
					sheet[f'B{row}'] = out_of_fortnight.id
					sheet[f'C{row}'] = i.employee_id.name
					sheet[f'D{row}'] = out_of_fortnight.name_input_id.code
					sheet[f'E{row}'] = out_of_fortnight.amount
					sheet[f'F{row}'] = out_of_fortnight.type
					row += 1
			else:
				sheet[f'A{row}'] = i.codigo_trabajador
				sheet[f'C{row}'] = i.employee_id.name
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])
