# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import *
import base64
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

class AccountAnexoWizard(models.TransientModel):
	_name = 'account.anexo.wizard'
	_description = 'Account Anexo Wizard'

	name = fields.Char()
	company_id = fields.Many2one('res.company',string=u'Compañia',required=True, default=lambda self: self.env.company,readonly=True)
	fiscal_year_id = fields.Many2one('account.fiscal.year',string='Ejercicio',required=True)
	period_from = fields.Many2one('account.period',string='Periodo Inicial',required=True)
	period_to = fields.Many2one('account.period',string='Periodo Final',required=True)

	@api.onchange('company_id')
	def get_fiscal_year(self):
		if self.company_id:
			fiscal_year = self.env['main.parameter'].search([('company_id','=',self.company_id.id)],limit=1).fiscal_year
			if fiscal_year:
				self.fiscal_year_id = fiscal_year.id
	
	def get_report(self):
		self.ensure_one()
		self.env.cr.execute(self._get_f1_balance_sql())
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = 'HT BALAN'
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		
		self.env.cr.execute(self._get_f1_register_sql())
		resx = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		resx.insert(0, colnames)
		wsx = wb.create_sheet('HT REG')
		row_position = 1
		col_position = 1
		for index, row in enumerate(resx, row_position):
			for col, val in enumerate(row, col_position):
				wsx.cell(row=index, column=col).value = val

		##
		self.env.cr.execute(self._get_sql_10())
		resx = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		resx.insert(0, colnames)
		wsx = wb.create_sheet('10')
		row_position = 1
		col_position = 1
		for index, row in enumerate(resx, row_position):
			for col, val in enumerate(row, col_position):
				wsx.cell(row=index, column=col).value = val
		##
		code = ['12','13','14','16','17','37','41','42','43','44','45','46','47','48','49']
		for cod in code:
			self.env.cr.execute(self.sql_saldos(cod))
			resx = self.env.cr.fetchall()
			colnames = [
				desc[0] for desc in self.env.cr.description
			]
			resx.insert(0, colnames)
			wsx = wb.create_sheet(cod)
			row_position = 1
			col_position = 1
			for index, row in enumerate(resx, row_position):
				for col, val in enumerate(row, col_position):
					wsx.cell(row=index, column=col).value = val
		##
					
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('Anexos.xlsx',output_datas)
	
	def _get_sql_10(self):
		sql = """
			SELECT * FROM (
			SELECT cuenta, nomenclatura, activo - pasivo as saldo
			FROM get_f1_register('{period_from}','{period_to}',{company},'pen')
			where left(cuenta,2) = '10' and activo - pasivo <> 0
			UNION ALL
			SELECT 
			null::character varying as cuenta,
			'SUMAS'::text as nomenclatura,
			sum(activo) - sum(pasivo) as saldo
			FROM get_f1_register('{period_from}','{period_to}',{company},'pen')
			where left(cuenta,2) = '10'
				)T
		""".format(
				period_from = self.period_from.code,
				period_to = self.period_to.code,
				company = self.company_id.id
			)

		return sql

	def _get_f1_register_sql(self):
		sql = """
			SELECT * FROM (
			SELECT mayor, cuenta, nomenclatura, debe, haber, saldo_deudor, saldo_acreedor, activo, pasivo, perdinat, ganannat, perdifun, gananfun, rubro
			FROM get_f1_register('{period_from}','{period_to}',{company},'pen')
			UNION ALL
			SELECT 
			null::text as mayor,
			null::character varying as cuenta,
			'SUMAS'::text as nomenclatura,
			sum(debe) as debe,
			sum(haber) as haber,
			sum(saldo_deudor) as saldo_deudor,
			sum(saldo_acreedor) as saldo_acreedor,
			sum(activo) as activo,
			sum(pasivo) as pasivo,
			sum(perdinat) as perdinat,
			sum(ganannat) as ganannat,
			sum(perdifun) as perdifun,
			sum(gananfun) as gananfun,
			null::text as rubro
			FROM get_f1_register('{period_from}','{period_to}',{company},'pen')
			UNION ALL
			SELECT 
			null::text as mayor,
			null::character varying as cuenta,
			'UTILIDAD O PERDIDA'::text as nomenclatura,
			case
				when sum(debe) < sum(haber)
				then sum(haber) - sum(debe)
				else 0
			end as debe,
			case
				when sum(debe) > sum(haber)
				then sum(debe) - sum(haber) 
				else 0
			end as haber,
			case
				when sum(saldo_deudor) < sum(saldo_acreedor)
				then sum(saldo_acreedor) - sum(saldo_deudor)
				else 0
			end as saldo_deudor,
			case
				when sum(saldo_deudor) > sum(saldo_acreedor)
				then sum(saldo_deudor) - sum(saldo_acreedor)
				else 0
			end as saldo_acreedor,
			case
				when sum(activo) < sum(pasivo)
				then sum(pasivo) - sum(activo)
				else 0
			end as activo,
			case
				when sum(activo) > sum(pasivo)
				then sum(activo) - sum(pasivo)
				else 0
			end as pasivo,
			case
				when sum(perdinat) < sum(ganannat)
				then sum(ganannat) - sum(perdinat)
				else 0
			end as perdinat,
			case
				when sum(perdinat) > sum(ganannat)
				then sum(perdinat) - sum(ganannat)
				else 0
			end as ganannat,
			case
				when sum(perdifun) < sum(gananfun)
				then sum(gananfun) - sum(perdifun)
				else 0
			end as perdifun,
			case
				when sum(perdifun) > sum(gananfun)
				then sum(perdifun) - sum(gananfun)
				else 0
			end as gananfun,
			null::text as rubro
			FROM get_f1_register('{period_from}','{period_to}',{company},'pen')
				)T
		""".format(
				period_from = self.period_from.code,
				period_to = self.period_to.code,
				company = self.company_id.id
			)

		return sql
		
	def _get_f1_balance_sql(self):
		sql = """
			SELECT * FROM (
			SELECT *
			FROM get_f1_balance('{period_from}','{period_to}',{company},'pen')
			UNION ALL
			SELECT 
			null::text as mayor,
			'SUMAS'::text as nomenclatura,
			sum(debe) as debe,
			sum(haber) as haber,
			sum(saldo_deudor) as saldo_deudor,
			sum(saldo_acreedor) as saldo_acreedor,
			sum(activo) as activo,
			sum(pasivo) as pasivo,
			sum(perdinat) as perdinat,
			sum(ganannat) as ganannat,
			sum(perdifun) as perdifun,
			sum(gananfun) as gananfun
			FROM get_f1_balance('{period_from}','{period_to}',{company},'pen')
			UNION ALL
			SELECT 
			null::text as mayor,
			'UTILIDAD O PERDIDA'::text as nomenclatura,
			case
				when sum(debe) < sum(haber)
				then sum(haber) - sum(debe)
				else 0
			end as debe,
			case
				when sum(debe) > sum(haber)
				then sum(debe) - sum(haber) 
				else 0
			end as haber,
			case
				when sum(saldo_deudor) < sum(saldo_acreedor)
				then sum(saldo_acreedor) - sum(saldo_deudor)
				else 0
			end as saldo_deudor,
			case
				when sum(saldo_deudor) > sum(saldo_acreedor)
				then sum(saldo_deudor) - sum(saldo_acreedor)
				else 0
			end as saldo_acreedor,
			case
				when sum(activo) < sum(pasivo)
				then sum(pasivo) - sum(activo)
				else 0
			end as activo,
			case
				when sum(activo) > sum(pasivo)
				then sum(activo) - sum(pasivo)
				else 0
			end as pasivo,
			case
				when sum(perdinat) < sum(ganannat)
				then sum(ganannat) - sum(perdinat)
				else 0
			end as perdinat,
			case
				when sum(perdinat) > sum(ganannat)
				then sum(perdinat) - sum(ganannat)
				else 0
			end as ganannat,
			case
				when sum(perdifun) < sum(gananfun)
				then sum(gananfun) - sum(perdifun)
				else 0
			end as perdifun,
			case
				when sum(perdifun) > sum(gananfun)
				then sum(perdifun) - sum(gananfun)
				else 0
			end as gananfun
			FROM get_f1_balance('{period_from}','{period_to}',{company},'pen')
				)T
		""".format(
				period_from = self.period_from.code,
				period_to = self.period_to.code,
				company = self.company_id.id
			)
		return sql
	
	def sql_saldos(self,code):
		sql = """
			SELECT periodo, fecha_con, libro,voucher,td_partner,doc_partner,partner,td_sunat,nro_comprobante,
			fecha_doc,fecha_ven,cuenta,moneda,debe,haber,saldo_mn,saldo_me
			FROM get_saldos_anexos('{date_from}','{date_to}',{company},'{code}')
		""".format(
				date_from = self.period_from.date_start.strftime('%Y/%m/%d'),
				date_to = self.period_to.date_end.strftime('%Y/%m/%d'),
				company = self.company_id.id,
				code = code
			)
		return sql